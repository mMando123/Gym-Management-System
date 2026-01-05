"""Members module frame for Gym Management System.

Provides:
- Listing/searching/filtering/pagination
- Add/Edit/View/Delete (soft delete)
- Export (XLSX if openpyxl available, otherwise CSV)

Designed for RTL Arabic UI.
"""

from __future__ import annotations

import re
import tkinter as tk
from datetime import date, datetime
from tkinter import filedialog, messagebox
from tkinter import ttk
from typing import Any, Optional

import ttkbootstrap as tb

try:
    from ttkbootstrap.tableview import Tableview  # type: ignore
except Exception:
    Tableview = None  # type: ignore

from PIL import Image, ImageTk

import config
from database import DatabaseManager


class MembersFrame(tb.Frame):
    """Frame that manages members: list/add/edit/delete/search."""

    def __init__(self, parent: ttk.Widget, db: DatabaseManager | None, user_data: dict | None = None) -> None:
        super().__init__(parent)
        self.db = db
        self.user_data = user_data or {}

        self.breakpoint: str = "desktop"
        self._selected_member_id: int | None = None

        self.search_var = tk.StringVar(master=self, value="")
        self.status_var = tk.StringVar(master=self, value="الكل")
        self.sub_type_var = tk.StringVar(master=self, value="الكل")

        self.current_page: int = 1
        self.items_per_page: int = 25
        self.total_pages: int = 1
        self.per_page_var = tk.StringVar(master=self, value="25")

        self._all_rows: list[dict[str, Any]] = []
        self._filtered_rows: list[dict[str, Any]] = []
        self._sort_col: str | None = None
        self._sort_desc: bool = False

        self._photo_cache: dict[str, ImageTk.PhotoImage] = {}

        self.configure(padding=10)
        self.create_toolbar()
        self.create_filters()
        self.create_table()
        self.create_pagination()
        self.create_status_bar()

        self.refresh_data()

    def on_breakpoint_change(self, breakpoint: str) -> None:
        self.breakpoint = breakpoint
        self._apply_responsive_layout()
        try:
            self._layout_filter_controls()
        except Exception:
            pass
        self._render_cards()

    # ------------------------------
    # UI
    # ------------------------------

    def create_toolbar(self) -> None:
        toolbar = tb.Frame(self)
        toolbar.pack(fill="x", pady=(0, 10))

        tb.Button(toolbar, text="➕ إضافة عضو جديد", bootstyle="success", command=self.show_add_member_dialog).pack(
            side="right", padx=4
        )
        self.btn_activate = tb.Button(toolbar, text="✅ تنشيط", bootstyle="success-outline", command=self.activate_selected_member)
        self.btn_activate.pack(side="right", padx=4)
        self.btn_activate.configure(state="disabled")
        tb.Button(toolbar, text="✅ تنشيط الكل", bootstyle="success-outline", command=self.activate_all_inactive_members).pack(
            side="right", padx=4
        )
        tb.Button(toolbar, text="✏️ تعديل", bootstyle="info", command=self.edit_selected_member).pack(side="right", padx=4)
        tb.Button(toolbar, text="🗑️ حذف", bootstyle="danger", command=self.delete_selected_member).pack(side="right", padx=4)
        tb.Button(toolbar, text="👁️ عرض", bootstyle="secondary", command=self.view_member_details).pack(side="right", padx=4)
        tb.Button(toolbar, text="📤 تصدير", bootstyle="secondary", command=self.export_to_excel).pack(side="right", padx=4)
        tb.Button(toolbar, text="🔄 تحديث", bootstyle="secondary", command=self.refresh_data).pack(side="left", padx=4)

    def create_filters(self) -> None:
        filter_frame = tb.Frame(self)
        filter_frame.pack(fill="x", pady=(0, 10))

        self.filter_frame = filter_frame

        self.lbl_search_icon = tb.Label(filter_frame, text="🔍", font=("Cairo", 12))
        self.search_entry = tb.Entry(filter_frame, textvariable=self.search_var, justify="right")
        self.search_entry.insert(0, "بحث بالاسم أو رقم الهاتف أو رقم العضوية...")

        def on_focus_in(_e):
            if self.search_entry.get().strip() == "بحث بالاسم أو رقم الهاتف أو رقم العضوية...":
                self.search_entry.delete(0, tk.END)

        def on_focus_out(_e):
            if not self.search_entry.get().strip():
                self.search_entry.insert(0, "بحث بالاسم أو رقم الهاتف أو رقم العضوية...")

        self.search_entry.bind("<FocusIn>", on_focus_in)
        self.search_entry.bind("<FocusOut>", on_focus_out)

        self.search_var.trace_add("write", lambda *_: self.on_search_change())

        self.lbl_status = tb.Label(filter_frame, text="الحالة:", font=("Cairo", 10, "bold"))
        status_values = ["الكل", "نشط", "غير نشط", "مجمد"]
        self.status_combo = tb.Combobox(
            filter_frame,
            textvariable=self.status_var,
            values=status_values,
            state="readonly",
            width=12,
            justify="right",
        )
        self.status_combo.bind("<<ComboboxSelected>>", lambda _e: self.apply_filters())

        self.lbl_sub = tb.Label(filter_frame, text="نوع الاشتراك:", font=("Cairo", 10, "bold"))
        sub_values = ["الكل"] + [str(x.get("name_ar")) for x in config.SUBSCRIPTION_TYPES]
        self.sub_combo = tb.Combobox(
            filter_frame,
            textvariable=self.sub_type_var,
            values=sub_values,
            state="readonly",
            width=14,
            justify="right",
        )
        self.sub_combo.bind("<<ComboboxSelected>>", lambda _e: self.apply_filters())

        self.btn_clear_filters = tb.Button(
            filter_frame,
            text="مسح الفلاتر",
            bootstyle="secondary-outline",
            command=self.clear_filters,
        )

        self._layout_filter_controls()

    def _layout_filter_controls(self) -> None:
        if not hasattr(self, "filter_frame"):
            return

        widgets = [
            self.lbl_search_icon,
            self.search_entry,
            self.lbl_status,
            self.status_combo,
            self.lbl_sub,
            self.sub_combo,
            self.btn_clear_filters,
        ]

        for w in widgets:
            try:
                w.grid_forget()
            except Exception:
                pass

        for i in range(10):
            try:
                self.filter_frame.columnconfigure(i, weight=0)
            except Exception:
                pass

        if getattr(self, "breakpoint", "desktop") == "mobile":
            self.filter_frame.columnconfigure(0, weight=1)
            self.lbl_search_icon.grid(row=0, column=1, sticky="e", padx=(0, 6), pady=3)
            self.search_entry.grid(row=0, column=0, sticky="ew", pady=3, ipady=4)

            self.lbl_status.grid(row=1, column=1, sticky="e", padx=(0, 6), pady=3)
            self.status_combo.grid(row=1, column=0, sticky="ew", pady=3)

            self.lbl_sub.grid(row=2, column=1, sticky="e", padx=(0, 6), pady=3)
            self.sub_combo.grid(row=2, column=0, sticky="ew", pady=3)

            self.btn_clear_filters.grid(row=3, column=0, columnspan=2, sticky="ew", pady=(6, 0))
            return

        self.filter_frame.columnconfigure(0, weight=1)
        self.lbl_search_icon.grid(row=0, column=7, sticky="e", padx=(0, 6), pady=3)
        self.search_entry.grid(row=0, column=6, sticky="ew", padx=(0, 12), pady=3, ipady=4)

        self.lbl_status.grid(row=0, column=5, sticky="e", padx=(0, 6), pady=3)
        self.status_combo.grid(row=0, column=4, sticky="w", padx=(0, 12), pady=3)

        self.lbl_sub.grid(row=0, column=3, sticky="e", padx=(0, 6), pady=3)
        self.sub_combo.grid(row=0, column=2, sticky="w", padx=(0, 12), pady=3)

        self.btn_clear_filters.grid(row=0, column=0, sticky="w", padx=(12, 0), pady=3)

    def create_table(self) -> None:
        table_frame = tb.Frame(self)
        table_frame.pack(fill="both", expand=True)
        self.table_wrap = table_frame

        columns = (
            "db_id",
            "member_code",
            "photo",
            "name",
            "phone",
            "gender",
            "subscription",
            "end_date",
            "remaining",
            "status",
        )

        self.tree = ttk.Treeview(table_frame, columns=columns, show="headings", selectmode="extended")

        column_config = {
            "db_id": {"text": "#", "width": 0, "anchor": "center"},
            "member_code": {"text": "رقم العضوية", "width": 110, "anchor": "center"},
            "photo": {"text": "الصورة", "width": 70, "anchor": "center"},
            "name": {"text": "الاسم الكامل", "width": 220, "anchor": "e"},
            "phone": {"text": "رقم الهاتف", "width": 120, "anchor": "center"},
            "gender": {"text": "الجنس", "width": 80, "anchor": "center"},
            "subscription": {"text": "نوع الاشتراك", "width": 120, "anchor": "center"},
            "end_date": {"text": "تاريخ الانتهاء", "width": 120, "anchor": "center"},
            "remaining": {"text": "المتبقي", "width": 80, "anchor": "center"},
            "status": {"text": "الحالة", "width": 90, "anchor": "center"},
        }

        for col, cfg in column_config.items():
            self.tree.heading(col, text=cfg["text"], command=lambda c=col: self.sort_by_column(c))
            self.tree.column(col, width=cfg["width"], minwidth=cfg["width"], anchor=cfg["anchor"], stretch=False)

        # Hide db id column
        self.tree.column("db_id", width=0, stretch=False)

        self.tree_vsb = ttk.Scrollbar(table_frame, orient="vertical", command=self.tree.yview)
        self.tree_hsb = ttk.Scrollbar(table_frame, orient="horizontal", command=self.tree.xview)
        self.tree.configure(yscrollcommand=self.tree_vsb.set, xscrollcommand=self.tree_hsb.set)

        self.tree.pack(side="left", fill="both", expand=True)
        self.tree_vsb.pack(side="right", fill="y")
        self.tree_hsb.pack(side="bottom", fill="x")

        self.cards_canvas = tk.Canvas(table_frame, highlightthickness=0)
        self.cards_vsb = ttk.Scrollbar(table_frame, orient="vertical", command=self.cards_canvas.yview)
        self.cards_canvas.configure(yscrollcommand=self.cards_vsb.set)

        self.cards_inner = tb.Frame(self.cards_canvas)
        self.cards_window = self.cards_canvas.create_window((0, 0), window=self.cards_inner, anchor="nw")

        self.cards_inner.bind(
            "<Configure>",
            lambda _e: self.cards_canvas.configure(scrollregion=self.cards_canvas.bbox("all")),
        )
        self.cards_canvas.bind(
            "<Configure>",
            lambda e: self.cards_canvas.itemconfigure(self.cards_window, width=e.width),
        )

        self.tree.bind("<Double-1>", lambda _e: self.view_member_details())
        self.tree.bind("<Delete>", lambda _e: self.delete_selected_member())
        self.tree.bind("<Return>", lambda _e: self.view_member_details())
        self.tree.bind("<<TreeviewSelect>>", lambda _e: self.on_selection_change())
        self.tree.bind("<Button-3>", self.show_context_menu)

        self.tree.tag_configure("active", background="#dcfce7")
        self.tree.tag_configure("expiring", background="#fef3c7")
        self.tree.tag_configure("expired", background="#fee2e2")
        self.tree.tag_configure("inactive", background="#e5e7eb")

        self.context_menu = tk.Menu(self, tearoff=0)
        self.context_menu.add_command(label="👁️ عرض", command=self.view_member_details)
        self.context_menu.add_command(label="✏️ تعديل", command=self.edit_selected_member)
        self.context_menu.add_command(label="✅ تنشيط", command=self.activate_selected_member)
        self.context_menu.add_separator()
        self.context_menu.add_command(label="🗑️ حذف", command=self.delete_selected_member)

        self._apply_responsive_layout()

    def _apply_responsive_layout(self) -> None:
        try:
            if self.breakpoint == "mobile":
                try:
                    self.tree.pack_forget()
                    self.tree_vsb.pack_forget()
                    self.tree_hsb.pack_forget()
                except Exception:
                    pass

                self.cards_canvas.pack(side="left", fill="both", expand=True)
                self.cards_vsb.pack(side="right", fill="y")
            else:
                try:
                    self.cards_canvas.pack_forget()
                    self.cards_vsb.pack_forget()
                except Exception:
                    pass

                self.tree.pack(side="left", fill="both", expand=True)
                self.tree_vsb.pack(side="right", fill="y")
                self.tree_hsb.pack(side="bottom", fill="x")
        except Exception:
            pass

    def create_pagination(self) -> None:
        pagination_frame = tb.Frame(self)
        pagination_frame.pack(fill="x", pady=(10, 6))

        self.btn_prev = tb.Button(pagination_frame, text="◀ السابق", bootstyle="secondary", command=self.prev_page)
        self.btn_prev.pack(side="right", padx=4)

        self.page_label = tb.Label(pagination_frame, text="صفحة 1 من 1", font=("Cairo", 10))
        self.page_label.pack(side="right", padx=10)

        self.btn_next = tb.Button(pagination_frame, text="التالي ▶", bootstyle="secondary", command=self.next_page)
        self.btn_next.pack(side="right", padx=4)

        tb.Label(pagination_frame, text="عرض:", font=("Cairo", 10, "bold")).pack(side="left", padx=(0, 6))
        per_page_combo = tb.Combobox(
            pagination_frame,
            textvariable=self.per_page_var,
            values=["10", "25", "50", "100"],
            state="readonly",
            width=6,
        )
        per_page_combo.pack(side="left")
        per_page_combo.bind("<<ComboboxSelected>>", lambda _e: self.on_per_page_change())

    def create_status_bar(self) -> None:
        status_frame = tb.Frame(self)
        status_frame.pack(fill="x")

        self.stats_label = tb.Label(status_frame, text="إجمالي الأعضاء: 0 │ نشط: 0 │ غير نشط: 0 │ مجمد: 0", font=("Cairo", 10))
        self.stats_label.pack(side="right")

    # ------------------------------
    # Data
    # ------------------------------

    def refresh_data(self) -> None:
        """Reload members list from the database."""

        if self.db is None:
            messagebox.showerror("خطأ", "قاعدة البيانات غير جاهزة")
            return

        try:
            self._all_rows = self._fetch_members()
            self.apply_filters(reset_page=True)
        except Exception as e:
            messagebox.showerror("خطأ", f"تعذر تحميل بيانات الأعضاء: {e}")

    def _fetch_members(self) -> list[dict[str, Any]]:
        """Fetch members with latest subscription info."""

        query = """
            SELECT m.*,
                   s.end_date AS sub_end_date,
                   s.status AS sub_status,
                   st.name_ar AS sub_name_ar
            FROM members m
            LEFT JOIN (
                SELECT s1.*
                FROM subscriptions s1
                JOIN (
                    SELECT member_id, MAX(date(end_date)) AS max_end
                    FROM subscriptions
                    GROUP BY member_id
                ) s2 ON s1.member_id = s2.member_id AND date(s1.end_date) = s2.max_end
            ) s ON s.member_id = m.id
            LEFT JOIN subscription_types st ON st.id = s.subscription_type_id
            ORDER BY m.id DESC
        """

        with self.db.get_connection() as conn:
            rows = conn.execute(query).fetchall()
            return [dict(r) for r in rows]

    def apply_filters(self, reset_page: bool = False) -> None:
        """Apply current search/filters to the in-memory list."""

        if reset_page:
            self.current_page = 1

        q = self.search_var.get().strip()
        if q == "بحث بالاسم أو رقم الهاتف أو رقم العضوية...":
            q = ""

        status_ar = self.status_var.get().strip()
        sub_ar = self.sub_type_var.get().strip()

        def match(row: dict[str, Any]) -> bool:
            if q:
                hay = " ".join(
                    [
                        str(row.get("member_code", "")),
                        str(row.get("first_name", "")),
                        str(row.get("last_name", "")),
                        str(row.get("phone", "")),
                    ]
                )
                if q.lower() not in hay.lower():
                    return False

            if status_ar and status_ar != "الكل":
                status_map = {
                    "نشط": "active",
                    "غير نشط": "inactive",
                    "مجمد": "frozen",
                }
                expected = status_map.get(status_ar)
                if expected and str(row.get("status", "")) != expected:
                    return False

            if sub_ar and sub_ar != "الكل":
                if str(row.get("sub_name_ar") or "") != sub_ar:
                    return False

            return True

        self._filtered_rows = [r for r in self._all_rows if match(r)]

        if self._sort_col:
            self._filtered_rows = self._sorted_rows(self._filtered_rows, self._sort_col, self._sort_desc)

        self._update_pagination()
        self._render_page()
        self._update_status_stats()

    def _sorted_rows(self, rows: list[dict[str, Any]], col: str, desc: bool) -> list[dict[str, Any]]:
        def key_fn(r: dict[str, Any]):
            v = r.get(col)
            return "" if v is None else v

        return sorted(rows, key=key_fn, reverse=desc)

    def _update_pagination(self) -> None:
        self.items_per_page = int(self.per_page_var.get() or "25")
        total_items = len(self._filtered_rows)
        self.total_pages = max((total_items + self.items_per_page - 1) // self.items_per_page, 1)
        self.current_page = max(min(self.current_page, self.total_pages), 1)
        self.page_label.configure(text=f"صفحة {self.current_page} من {self.total_pages}")

        self.btn_prev.configure(state=("normal" if self.current_page > 1 else "disabled"))
        self.btn_next.configure(state=("normal" if self.current_page < self.total_pages else "disabled"))

    def _render_page(self) -> None:
        for iid in self.tree.get_children():
            self.tree.delete(iid)

        start = (self.current_page - 1) * self.items_per_page
        end = start + self.items_per_page
        page_rows = self._filtered_rows[start:end]

        today = date.today()

        for r in page_rows:
            member_id = int(r.get("id"))
            member_code = str(r.get("member_code") or "")
            name = f"{r.get('first_name', '')} {r.get('last_name', '')}".strip()
            phone = str(r.get("phone") or "")

            gender = str(r.get("gender") or "")
            if gender == "male":
                gender_ar = "ذكر"
            elif gender == "female":
                gender_ar = "أنثى"
            else:
                gender_ar = "-"

            sub_name = str(r.get("sub_name_ar") or "-")
            end_date = str(r.get("sub_end_date") or "-")

            remaining = "-"
            tag = "active"

            member_status = str(r.get("status") or "active")
            if member_status != "active":
                tag = "inactive"

            if end_date and end_date != "-":
                try:
                    d_end = datetime.strptime(end_date, config.DATE_FORMAT).date()
                    delta = (d_end - today).days
                    remaining = f"{delta}"
                    if delta < 0:
                        tag = "expired"
                    elif delta <= 7:
                        tag = "expiring"
                except Exception:
                    remaining = "-"

            status_ar = {
                "active": "نشط",
                "inactive": "غير نشط",
                "frozen": "مجمد",
            }.get(member_status, member_status)

            self.tree.insert(
                "",
                "end",
                iid=str(member_id),
                values=(member_id, member_code, "👤", name, phone, gender_ar, sub_name, end_date, remaining, status_ar),
                tags=(tag,),
            )

        self._render_cards(page_rows)

    def _render_cards(self, page_rows: list[dict[str, Any]] | None = None) -> None:
        if getattr(self, "breakpoint", "desktop") != "mobile":
            return
        if not hasattr(self, "cards_inner"):
            return

        for child in self.cards_inner.winfo_children():
            child.destroy()

        if page_rows is None:
            start = (self.current_page - 1) * self.items_per_page
            end = start + self.items_per_page
            page_rows = self._filtered_rows[start:end]

        today = date.today()

        for r in page_rows:
            member_id = int(r.get("id"))
            member_code = str(r.get("member_code") or "")
            name = f"{r.get('first_name', '')} {r.get('last_name', '')}".strip() or "-"
            phone = str(r.get("phone") or "-")

            sub_name = str(r.get("sub_name_ar") or "-")
            end_date = str(r.get("sub_end_date") or "-")

            remaining = "-"
            tag = "active"
            member_status = str(r.get("status") or "active")
            if member_status != "active":
                tag = "inactive"

            if end_date and end_date != "-":
                try:
                    d_end = datetime.strptime(end_date, config.DATE_FORMAT).date()
                    delta = (d_end - today).days
                    remaining = f"{delta}"
                    if delta < 0:
                        tag = "expired"
                    elif delta <= 7:
                        tag = "expiring"
                except Exception:
                    remaining = "-"

            status_ar = {
                "active": "نشط",
                "inactive": "غير نشط",
                "frozen": "مجمد",
            }.get(member_status, member_status)

            card = tb.Frame(self.cards_inner, padding=10, bootstyle="secondary")
            card.pack(fill="x", pady=6)
            card.configure(cursor="hand2")

            top = tb.Frame(card)
            top.pack(fill="x")
            tb.Label(top, text=member_code, font=("Cairo", 11, "bold"), anchor="e").pack(side="right")
            tb.Label(top, text=status_ar, font=("Cairo", 10, "bold"), anchor="w").pack(side="left")

            tb.Label(card, text=name, font=("Cairo", 13, "bold"), anchor="e").pack(fill="x", pady=(6, 0))
            tb.Label(card, text=f"هاتف: {phone}", font=("Cairo", 10), anchor="e").pack(fill="x")
            tb.Label(card, text=f"اشتراك: {sub_name}", font=("Cairo", 10), anchor="e").pack(fill="x")
            tb.Label(card, text=f"ينتهي: {end_date} | المتبقي: {remaining}", font=("Cairo", 10), anchor="e").pack(
                fill="x", pady=(2, 0)
            )

            def open_member(_e=None, mid=member_id):
                self._selected_member_id = int(mid)
                try:
                    self.tree.selection_set(str(mid))
                except Exception:
                    pass
                self.view_member_details()

            for w in (card, top):
                w.bind("<Button-1>", open_member)

    def _update_status_stats(self) -> None:
        total = len(self._filtered_rows)
        active = sum(1 for r in self._filtered_rows if str(r.get("status")) == "active")
        inactive = sum(1 for r in self._filtered_rows if str(r.get("status")) == "inactive")
        frozen = sum(1 for r in self._filtered_rows if str(r.get("status")) == "frozen")
        self.stats_label.configure(text=f"إجمالي الأعضاء: {total} │ نشط: {active} │ غير نشط: {inactive} │ مجمد: {frozen}")

    # ------------------------------
    # Events
    # ------------------------------

    def on_search_change(self) -> None:
        self.apply_filters(reset_page=True)

    def clear_filters(self) -> None:
        self.search_var.set("")
        self.status_var.set("الكل")
        self.sub_type_var.set("الكل")
        self.apply_filters(reset_page=True)

    def on_selection_change(self) -> None:
        try:
            member_id = self._get_selected_member_id()
            if member_id is None:
                self.btn_activate.configure(state="disabled")
                return

            status_key = None
            for r in self._all_rows:
                try:
                    if int(r.get("id")) == int(member_id):
                        status_key = str(r.get("status") or "active")
                        break
                except Exception:
                    continue

            if status_key == "inactive":
                self.btn_activate.configure(state="normal")
            else:
                self.btn_activate.configure(state="disabled")
        except Exception:
            try:
                self.btn_activate.configure(state="disabled")
            except Exception:
                pass

    def activate_selected_member(self) -> None:
        if self.db is None:
            messagebox.showerror("خطأ", "قاعدة البيانات غير جاهزة")
            return

        member_id = self._get_selected_member_id()
        if member_id is None:
            messagebox.showwarning("تنبيه", "يرجى اختيار عضو")
            return

        member = self.db.get_member_by_id(member_id)
        if not member:
            messagebox.showerror("خطأ", "لم يتم العثور على العضو")
            return

        if str(member.get("status")) != "inactive":
            messagebox.showinfo("تنبيه", "العضو بالفعل نشط")
            return

        if not messagebox.askyesno("تأكيد", "هل تريد تنشيط هذا العضو؟"):
            return

        ok, msg = self.db.activate_member(member_id)
        if ok:
            messagebox.showinfo("تم", "تم تنشيط العضو")
            self.refresh_data()
        else:
            messagebox.showerror("خطأ", msg)

    def activate_all_inactive_members(self) -> None:
        if self.db is None:
            messagebox.showerror("خطأ", "قاعدة البيانات غير جاهزة")
            return

        if not messagebox.askyesno("تأكيد", "هل تريد تنشيط جميع الأعضاء غير النشطين؟"):
            return

        ok, msg, count = self.db.activate_all_inactive_members()
        if ok:
            messagebox.showinfo("تم", f"تم تنشيط {count} عضو")
            self.refresh_data()
        else:
            messagebox.showerror("خطأ", msg)

    def sort_by_column(self, col: str) -> None:
        if col == "photo":
            return

        if self._sort_col == col:
            self._sort_desc = not self._sort_desc
        else:
            self._sort_col = col
            self._sort_desc = False

        self.apply_filters(reset_page=False)

    def show_context_menu(self, event) -> None:
        try:
            self.context_menu.tk_popup(event.x_root, event.y_root)
        finally:
            try:
                self.context_menu.grab_release()
            except Exception:
                pass

    def on_double_click(self, _event) -> None:
        self.view_member_details()

    # ------------------------------
    # Pagination
    # ------------------------------

    def prev_page(self) -> None:
        if self.current_page > 1:
            self.current_page -= 1
            self._update_pagination()
            self._render_page()

    def next_page(self) -> None:
        if self.current_page < self.total_pages:
            self.current_page += 1
            self._update_pagination()
            self._render_page()

    def on_per_page_change(self) -> None:
        self.current_page = 1
        self.apply_filters(reset_page=True)

    # ------------------------------
    # Actions
    # ------------------------------

    def _get_selected_member_id(self) -> int | None:
        sel = self.tree.selection()
        if not sel:
            if self._selected_member_id is not None:
                return self._selected_member_id
            return None
        values = self.tree.item(sel[0], "values")
        try:
            return int(values[0])
        except Exception:
            return None

    def show_add_member_dialog(self) -> None:
        if self.db is None:
            messagebox.showerror("خطأ", "قاعدة البيانات غير جاهزة")
            return

        dlg = MemberDialog(self.winfo_toplevel(), self.db, member_data=None)
        self.wait_window(dlg)
        if dlg.result:
            self.refresh_data()

    def edit_selected_member(self) -> None:
        if self.db is None:
            messagebox.showerror("خطأ", "قاعدة البيانات غير جاهزة")
            return

        member_id = self._get_selected_member_id()
        if member_id is None:
            messagebox.showwarning("تنبيه", "يرجى اختيار عضو")
            return

        member = self.db.get_member_by_id(member_id)
        if not member:
            messagebox.showerror("خطأ", "لم يتم العثور على العضو")
            return

        dlg = MemberDialog(self.winfo_toplevel(), self.db, member_data=member)
        self.wait_window(dlg)
        if dlg.result:
            self.refresh_data()

    def view_member_details(self) -> None:
        if self.db is None:
            messagebox.showerror("خطأ", "قاعدة البيانات غير جاهزة")
            return

        member_id = self._get_selected_member_id()
        if member_id is None:
            messagebox.showwarning("تنبيه", "يرجى اختيار عضو")
            return

        member = self.db.get_member_by_id(member_id)
        if not member:
            messagebox.showerror("خطأ", "لم يتم العثور على العضو")
            return

        dlg = MemberDialog(self.winfo_toplevel(), self.db, member_data=member, read_only=True)
        self.wait_window(dlg)

    def delete_selected_member(self) -> None:
        if self.db is None:
            messagebox.showerror("خطأ", "قاعدة البيانات غير جاهزة")
            return

        member_id = self._get_selected_member_id()
        if member_id is None:
            messagebox.showwarning("تنبيه", "يرجى اختيار عضو")
            return

        if not messagebox.askyesno("تأكيد", "هل تريد حذف (إيقاف) العضو المحدد؟"):
            return

        ok, msg = self.db.delete_member(member_id)
        if ok:
            messagebox.showinfo("تم", "تم تحديث حالة العضو إلى غير نشط")
            self.refresh_data()
        else:
            messagebox.showerror("خطأ", msg)

    def export_to_excel(self) -> None:
        """Export current filtered list to Excel (xlsx) or CSV."""

        if not self._filtered_rows:
            messagebox.showwarning("تنبيه", "لا توجد بيانات للتصدير")
            return

        file_path = filedialog.asksaveasfilename(
            title="تصدير",
            defaultextension=".xlsx",
            filetypes=[("Excel", "*.xlsx"), ("CSV", "*.csv")],
        )
        if not file_path:
            return

        try:
            if file_path.lower().endswith(".xlsx"):
                self._export_xlsx(file_path)
            else:
                self._export_csv(file_path)
            messagebox.showinfo("تم", "تم التصدير بنجاح")
        except Exception as e:
            messagebox.showerror("خطأ", f"فشل التصدير: {e}")

    def _export_csv(self, file_path: str) -> None:
        import csv

        headers = ["رقم العضوية", "الاسم", "الهاتف", "الجنس", "نوع الاشتراك", "تاريخ الانتهاء", "الحالة"]
        with open(file_path, "w", newline="", encoding="utf-8") as f:
            w = csv.writer(f)
            w.writerow(headers)
            for r in self._filtered_rows:
                name = f"{r.get('first_name', '')} {r.get('last_name', '')}".strip()
                gender = r.get("gender")
                gender_ar = "ذكر" if gender == "male" else ("أنثى" if gender == "female" else "-")
                status_ar = {"active": "نشط", "inactive": "غير نشط", "frozen": "مجمد"}.get(str(r.get("status")), "-")
                w.writerow(
                    [
                        r.get("member_code", ""),
                        name,
                        r.get("phone", ""),
                        gender_ar,
                        r.get("sub_name_ar") or "-",
                        r.get("sub_end_date") or "-",
                        status_ar,
                    ]
                )

    def _export_xlsx(self, file_path: str) -> None:
        try:
            import openpyxl  # type: ignore
            from openpyxl.styles import Alignment, Font  # type: ignore
        except Exception as e:
            raise RuntimeError("openpyxl غير مثبت. اختر CSV أو قم بتثبيت openpyxl") from e

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Members"

        headers = ["رقم العضوية", "الاسم", "الهاتف", "الجنس", "نوع الاشتراك", "تاريخ الانتهاء", "الحالة"]
        ws.append(headers)

        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="right")

        for r in self._filtered_rows:
            name = f"{r.get('first_name', '')} {r.get('last_name', '')}".strip()
            gender = r.get("gender")
            gender_ar = "ذكر" if gender == "male" else ("أنثى" if gender == "female" else "-")
            status_ar = {"active": "نشط", "inactive": "غير نشط", "frozen": "مجمد"}.get(str(r.get("status")), "-")
            ws.append(
                [
                    r.get("member_code", ""),
                    name,
                    r.get("phone", ""),
                    gender_ar,
                    r.get("sub_name_ar") or "-",
                    r.get("sub_end_date") or "-",
                    status_ar,
                ]
            )

        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(horizontal="right")

        wb.save(file_path)


class MemberDialog(tk.Toplevel):
    """Dialog for adding/editing/viewing a member."""

    def __init__(
        self,
        parent: tk.Misc,
        db: DatabaseManager,
        member_data: dict[str, Any] | None = None,
        read_only: bool = False,
    ) -> None:
        super().__init__(parent)
        self.db = db
        self.member_data = member_data
        self.read_only = read_only

        self.result: dict[str, Any] | None = None
        self.photo_path: str | None = (member_data.get("photo_path") if member_data else None)
        self._photo_preview: ImageTk.PhotoImage | None = None

        self.title("عرض بيانات العضو" if read_only else ("تعديل بيانات العضو" if member_data else "إضافة عضو جديد"))
        self.geometry("720x720")
        self.minsize(420, 520)
        self.resizable(True, True)
        self.grab_set()

        self.vars: dict[str, tk.Variable] = {
            "member_code": tk.StringVar(master=self, value=str(member_data.get("member_code", "")) if member_data else ""),
            "first_name": tk.StringVar(master=self, value=str(member_data.get("first_name", "")) if member_data else ""),
            "last_name": tk.StringVar(master=self, value=str(member_data.get("last_name", "")) if member_data else ""),
            "phone": tk.StringVar(master=self, value=str(member_data.get("phone", "")) if member_data else ""),
            "email": tk.StringVar(master=self, value=str(member_data.get("email", "")) if member_data else ""),
            "gender": tk.StringVar(master=self, value=str(member_data.get("gender", "")) if member_data else ""),
            "date_of_birth": tk.StringVar(master=self, value=str(member_data.get("date_of_birth", "")) if member_data else ""),
            "national_id": tk.StringVar(master=self, value=str(member_data.get("national_id", "")) if member_data else ""),
            "address": tk.StringVar(master=self, value=str(member_data.get("address", "")) if member_data else ""),
            "emergency_contact": tk.StringVar(master=self, value=str(member_data.get("emergency_contact", "")) if member_data else ""),
            "emergency_phone": tk.StringVar(master=self, value=str(member_data.get("emergency_phone", "")) if member_data else ""),
            "notes": tk.StringVar(master=self, value=str(member_data.get("notes", "")) if member_data else ""),
            "status": tk.StringVar(master=self, value=str(member_data.get("status", "active")) if member_data else "active"),
        }

        self.create_widgets()
        self._center_window()

        self.bind("<Escape>", lambda _e: self.destroy())

        if self.read_only:
            self._set_read_only()

    def create_widgets(self) -> None:
        container = tb.Frame(self, padding=14)
        container.pack(fill="both", expand=True)

        notebook = ttk.Notebook(container)
        notebook.pack(fill="both", expand=True, padx=4, pady=4)

        tab_basic = tb.Frame(notebook, padding=16)
        tab_additional = tb.Frame(notebook, padding=16)

        notebook.add(tab_basic, text="البيانات الأساسية")
        notebook.add(tab_additional, text="بيانات إضافية")

        # Photo section
        photo_frame = tb.Labelframe(tab_basic, text="صورة العضو", padding=10)
        photo_frame.pack(fill="x", pady=(0, 12))

        self.photo_label = tb.Label(photo_frame, text="👤", font=("Arial", 48), width=10, anchor="center")
        self.photo_label.pack(side="right", padx=10)

        btns = tb.Frame(photo_frame)
        btns.pack(side="right", padx=10)

        tb.Button(btns, text="📷 اختيار صورة", command=self.select_photo).pack(fill="x", pady=2)
        tb.Button(btns, text="🗑️ حذف الصورة", bootstyle="danger-outline", command=self.remove_photo).pack(fill="x", pady=2)

        self._load_photo_preview()

        # Basic form
        form = tb.Frame(tab_basic)
        form.pack(fill="x")

        self._field_row(form, "الاسم الأول", self.vars["first_name"], 0)
        self._field_row(form, "اسم العائلة", self.vars["last_name"], 1)
        self._field_row(form, "رقم الهاتف", self.vars["phone"], 2)
        self._field_row(form, "البريد الإلكتروني", self.vars["email"], 3)

        gender_row = tb.Frame(form)
        gender_row.grid(row=4, column=0, sticky="ew", pady=6)
        gender_row.columnconfigure(0, weight=1)
        tb.Label(gender_row, text="الجنس", font=("Cairo", 10, "bold"), anchor="e").pack(side="right")
        gender_combo = tb.Combobox(
            gender_row,
            textvariable=self.vars["gender"],
            values=["", "male", "female"],
            state="readonly",
            width=18,
            justify="right",
        )
        gender_combo.pack(side="left")

        dob_row = tb.Frame(form)
        dob_row.grid(row=5, column=0, sticky="ew", pady=6)
        dob_row.columnconfigure(0, weight=1)
        tb.Label(dob_row, text="تاريخ الميلاد", font=("Cairo", 10, "bold"), anchor="e").pack(side="right")
        dob_entry = tb.DateEntry(dob_row, dateformat="%Y-%m-%d", firstweekday=6, bootstyle="secondary")
        dob_entry.pack(side="left", fill="x", expand=True)

        # Link DateEntry to our StringVar (DateEntry itself is a Frame; the entry is inside)
        try:
            dob_entry.entry.configure(textvariable=self.vars["date_of_birth"])  # type: ignore[attr-defined]
        except Exception:
            pass

        def _sync_dob_from_widget(_e=None) -> None:
            try:
                v = str(dob_entry.entry.get()).strip()  # type: ignore[attr-defined]
                self.vars["date_of_birth"].set(v)
            except Exception:
                pass

        try:
            dob_entry.bind("<<DateEntrySelected>>", _sync_dob_from_widget)
        except Exception:
            pass
        try:
            dob_entry.bind("<<CalendarSelected>>", _sync_dob_from_widget)
        except Exception:
            pass
        try:
            dob_entry.bind("<FocusOut>", _sync_dob_from_widget)
        except Exception:
            pass

        # Ensure initial value is compatible with DateEntry
        try:
            v = str(self.vars["date_of_birth"].get() or "").strip()
            if v:
                datetime.strptime(v, config.DATE_FORMAT)
        except Exception:
            self.vars["date_of_birth"].set("")
        self._field_row(form, "رقم الهوية", self.vars["national_id"], 6)

        # Additional form
        form2 = tb.Frame(tab_additional)
        form2.pack(fill="both", expand=True)

        self._field_row(form2, "العنوان", self.vars["address"], 0)
        self._field_row(form2, "اسم للطوارئ", self.vars["emergency_contact"], 1)
        self._field_row(form2, "هاتف للطوارئ", self.vars["emergency_phone"], 2)

        notes_frame = tb.Frame(form2)
        notes_frame.grid(row=3, column=0, sticky="nsew", pady=6)
        notes_frame.columnconfigure(0, weight=1)
        form2.rowconfigure(3, weight=1)
        tb.Label(notes_frame, text="ملاحظات", font=("Cairo", 10, "bold"), anchor="e").pack(anchor="e")
        self.notes_text = tk.Text(notes_frame, height=6, wrap="word")
        self.notes_text.pack(fill="both", expand=True)
        if self.vars["notes"].get():
            self.notes_text.insert("1.0", str(self.vars["notes"].get()))

        # Buttons
        btn_row = tb.Frame(container)
        btn_row.pack(fill="x", pady=(10, 0))

        if not self.read_only:
            tb.Button(btn_row, text="حفظ", bootstyle="success", command=self.on_save).pack(side="left", padx=6)
        tb.Button(btn_row, text="إغلاق", bootstyle="secondary", command=self.destroy).pack(side="left")

    def _field_row(self, parent: ttk.Widget, label: str, var: tk.Variable, row: int) -> None:
        r = tb.Frame(parent)
        r.grid(row=row, column=0, sticky="ew", pady=6)
        r.columnconfigure(0, weight=1)

        tb.Label(r, text=label, font=("Cairo", 10, "bold"), anchor="e").pack(side="right")
        e = tb.Entry(r, textvariable=var, justify="right")
        e.pack(side="left", fill="x", expand=True)

    def _set_read_only(self) -> None:
        for child in self.winfo_children():
            pass

        # Disable all entries/combos/text
        def disable_recursive(widget: tk.Misc) -> None:
            for w in widget.winfo_children():
                cls = w.winfo_class()
                try:
                    if cls in {"TEntry", "Entry", "TCombobox", "Combobox"}:
                        w.configure(state="disabled")
                    if isinstance(w, tk.Text):
                        w.configure(state="disabled")
                    if isinstance(w, (tb.Button, ttk.Button)):
                        if str(w.cget("text")) not in {"إغلاق"}:
                            w.configure(state="disabled")
                except Exception:
                    pass
                disable_recursive(w)

        disable_recursive(self)

    def _center_window(self) -> None:
        self.update_idletasks()
        w = self.winfo_width()
        h = self.winfo_height()
        sw = self.winfo_screenwidth()
        sh = self.winfo_screenheight()
        x = (sw - w) // 2
        y = (sh - h) // 2
        self.geometry(f"{w}x{h}+{x}+{y}")

    # ------------------------------
    # Photo
    # ------------------------------

    def select_photo(self) -> None:
        if self.read_only:
            return

        path = filedialog.askopenfilename(
            title="اختيار صورة",
            filetypes=[("Images", "*.png;*.jpg;*.jpeg;*.bmp")],
        )
        if not path:
            return
        self.photo_path = path
        self._load_photo_preview()

    def remove_photo(self) -> None:
        if self.read_only:
            return

        self.photo_path = None
        self._photo_preview = None
        self.photo_label.configure(image="", text="👤")

    def _load_photo_preview(self) -> None:
        if not self.photo_path:
            return

        try:
            img = Image.open(self.photo_path).convert("RGBA")
            img.thumbnail((140, 140))
            self._photo_preview = ImageTk.PhotoImage(img)
            self.photo_label.configure(image=self._photo_preview, text="")
        except Exception:
            self.photo_label.configure(image="", text="👤")

    # ------------------------------
    # Save / validation
    # ------------------------------

    def _validate(self) -> tuple[bool, str]:
        first = str(self.vars["first_name"].get()).strip()
        last = str(self.vars["last_name"].get()).strip()
        phone = str(self.vars["phone"].get()).strip()
        email = str(self.vars["email"].get()).strip()
        dob = str(self.vars["date_of_birth"].get()).strip()

        if not first:
            return False, "يرجى إدخال الاسم الأول"
        if not last:
            return False, "يرجى إدخال اسم العائلة"

        if not phone:
            return False, "يرجى إدخال رقم الهاتف"
        if not re.fullmatch(r"\+?\d{8,15}", phone):
            return False, "رقم الهاتف غير صحيح"

        if email and not re.fullmatch(r"[^@\s]+@[^@\s]+\.[^@\s]+", email):
            return False, "البريد الإلكتروني غير صحيح"

        if dob:
            try:
                datetime.strptime(dob, config.DATE_FORMAT)
            except Exception:
                return False, "صيغة تاريخ الميلاد يجب أن تكون YYYY-MM-DD"

        return True, ""

    def on_save(self) -> None:
        ok, msg = self._validate()
        if not ok:
            messagebox.showerror("خطأ", msg)
            return

        payload: dict[str, Any] = {
            "first_name": str(self.vars["first_name"].get()).strip(),
            "last_name": str(self.vars["last_name"].get()).strip(),
            "phone": str(self.vars["phone"].get()).strip(),
            "email": str(self.vars["email"].get()).strip() or None,
            "gender": str(self.vars["gender"].get()).strip() or None,
            "date_of_birth": str(self.vars["date_of_birth"].get()).strip() or None,
            "national_id": str(self.vars["national_id"].get()).strip() or None,
            "address": str(self.vars["address"].get()).strip() or None,
            "emergency_contact": str(self.vars["emergency_contact"].get()).strip() or None,
            "emergency_phone": str(self.vars["emergency_phone"].get()).strip() or None,
            "photo_path": self.photo_path,
            "notes": self.notes_text.get("1.0", "end").strip() or None,
        }

        try:
            if self.member_data:
                member_id = int(self.member_data.get("id"))
                success, message = self.db.update_member(member_id, **payload)
                if not success:
                    messagebox.showerror("خطأ", message)
                    return
                self.result = {"id": member_id}
                self.destroy()
            else:
                success, message, new_id = self.db.create_member(**payload)
                if not success:
                    messagebox.showerror("خطأ", message)
                    return
                self.result = {"id": new_id}
                self.destroy()
        except Exception as e:
            messagebox.showerror("خطأ", f"فشل الحفظ: {e}")


if __name__ == "__main__":
    root = tb.Window(themename="cosmo")
    root.title("MembersFrame Test")
    db = DatabaseManager()
    frame = MembersFrame(root, db, {"username": "admin", "role": "admin"})
    frame.pack(fill="both", expand=True)
    root.geometry("1100x650")
    root.mainloop()
