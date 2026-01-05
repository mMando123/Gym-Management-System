"""Subscriptions module frame for Gym Management System.

This module works with the existing database schema from database.py:
- subscription_types
- subscriptions
- payments

It provides:
- Summary cards (active/expiring/expired/monthly revenue)
- Filters (date range, subscription status, payment status)
- Subscriptions table with context menu
- New subscription and renewal dialogs

RTL Arabic UI.
"""

from __future__ import annotations

import tkinter as tk
from datetime import date, datetime, timedelta
from tkinter import messagebox
from tkinter import ttk
from typing import Any

import ttkbootstrap as tb
from ttkbootstrap.constants import *  # noqa: F403
from ttkbootstrap.dialogs import Messagebox
from ttkbootstrap.widgets import DateEntry, Meter

import config
from database import DatabaseManager
from utils import format_money, print_text_windows


# Local styling (kept here to avoid circular imports)
COLORS = {
    "primary": "#2563eb",
    "primary_dark": "#1e40af",
    "secondary": "#64748b",
    "success": "#059669",
    "warning": "#d97706",
    "danger": "#dc2626",
    "background": "#f1f5f9",
    "sidebar": "#1e293b",
    "sidebar_hover": "#334155",
    "sidebar_active": "#3b82f6",
    "text": "#1e293b",
    "text_light": "#64748b",
    "white": "#ffffff",
}

FONTS = {
    "heading": ("Cairo", 18, "bold"),
    "subheading": ("Cairo", 14, "bold"),
    "body": ("Cairo", 12),
    "small": ("Cairo", 10),
}


def _today_str() -> str:
    return date.today().isoformat()


def _month_start(d: date) -> date:
    return date(d.year, d.month, 1)


def _add_months(start: date, months: int) -> date:
    """Add months to a date while clamping day-of-month."""

    month = start.month - 1 + months
    year = start.year + month // 12
    month = month % 12 + 1

    if month == 12:
        next_month = date(year + 1, 1, 1)
    else:
        next_month = date(year, month + 1, 1)
    last_day = (next_month - timedelta(days=1)).day
    day = min(start.day, last_day)
    return date(year, month, day)


class SubscriptionsFrame(tb.Frame):
    """Frame to manage subscriptions."""

    def __init__(self, parent: ttk.Widget, db: DatabaseManager | None, user_data: dict | None = None) -> None:
        super().__init__(parent)
        self.db = db
        self.user_data = user_data or {}

        self.breakpoint: str = "desktop"
        self._selected_sub_id: int | None = None

        self.start_date_var = tk.StringVar(master=self, value="")
        self.end_date_var = tk.StringVar(master=self, value="")
        self.status_var = tk.StringVar(master=self, value="الكل")
        self.payment_status_var = tk.StringVar(master=self, value="الكل")

        self.summary_vars = {
            "active": tk.StringVar(master=self, value="0"),
            "expiring": tk.StringVar(master=self, value="0"),
            "expired": tk.StringVar(master=self, value="0"),
            "revenue": tk.StringVar(master=self, value="0"),
        }

        self._rows: list[dict[str, Any]] = []

        self.configure(padding=10)
        self.create_summary_cards()
        self.create_filter_bar()
        self.create_table()
        self.create_quick_actions()

        self.load_summary_cards()
        self.load_subscriptions_data()
        self.check_expiring_subscriptions()

        self._on_show_refreshing: bool = False
        try:
            self.bind("<Map>", lambda _e: self._on_show_refresh())
        except Exception:
            pass

    def _on_show_refresh(self) -> None:
        if self.db is None:
            return
        if getattr(self, "_on_show_refreshing", False):
            return
        self._on_show_refreshing = True

        def run() -> None:
            try:
                self.refresh()
            finally:
                self._on_show_refreshing = False

        try:
            self.after(120, run)
        except Exception:
            run()

    def on_breakpoint_change(self, breakpoint: str) -> None:
        self.breakpoint = breakpoint
        try:
            self._apply_responsive_layout()
        except Exception:
            pass
        try:
            self._layout_filter_controls()
        except Exception:
            pass
        self._render_cards()

    # ------------------------------
    # UI sections
    # ------------------------------

    def create_summary_cards(self) -> None:
        row = tb.Frame(self)
        row.pack(fill="x", pady=(0, 12))

        self._summary_card(row, "✅", "الاشتراكات النشطة", self.summary_vars["active"], "success")
        self._summary_card(row, "⏳", "تنتهي خلال 7 أيام", self.summary_vars["expiring"], "warning")
        self._summary_card(row, "⛔", "اشتراكات منتهية", self.summary_vars["expired"], "danger")
        self._summary_card(row, "💰", "إجمالي الإيرادات الشهرية", self.summary_vars["revenue"], "info")

    def _summary_card(self, parent: ttk.Widget, icon: str, title: str, value_var: tk.StringVar, style: str) -> None:
        card = tb.Frame(parent, padding=12, bootstyle=f"{style}")
        card.pack(side="right", padx=6, fill="x", expand=True)

        tb.Label(card, text=f"{icon}", font=("Cairo", 20, "bold"), bootstyle=f"inverse-{style}").pack(anchor="e")
        tb.Label(card, textvariable=value_var, font=("Cairo", 18, "bold"), bootstyle=f"inverse-{style}").pack(anchor="e")
        tb.Label(card, text=title, font=FONTS["small"], bootstyle=f"inverse-{style}").pack(anchor="e")

    def create_filter_bar(self) -> None:
        bar = tb.Frame(self)
        bar.pack(fill="x", pady=(0, 12))

        self.filter_controls = tb.Frame(bar)
        self.filter_controls.pack(fill="x")

        self.filter_actions = tb.Frame(bar)
        self.filter_actions.pack(fill="x", pady=(6, 0))

        self.lbl_start = tb.Label(self.filter_controls, text="من تاريخ", font=("Cairo", 10, "bold"))
        self.start_date_entry = DateEntry(self.filter_controls, width=12, bootstyle="info")

        self.lbl_end = tb.Label(self.filter_controls, text="إلى تاريخ", font=("Cairo", 10, "bold"))
        self.end_date_entry = DateEntry(self.filter_controls, width=12, bootstyle="info")

        self.lbl_status = tb.Label(self.filter_controls, text="حالة الاشتراك", font=("Cairo", 10, "bold"))
        self.status_combo = tb.Combobox(
            self.filter_controls,
            textvariable=self.status_var,
            values=["الكل", "نشط", "منتهي", "معلق", "ملغي"],
            state="readonly",
            width=10,
            justify="right",
        )

        self.lbl_pay = tb.Label(self.filter_controls, text="حالة الدفع", font=("Cairo", 10, "bold"))
        self.pay_combo = tb.Combobox(
            self.filter_controls,
            textvariable=self.payment_status_var,
            values=["الكل", "مدفوع", "جزئي", "غير مدفوع"],
            state="readonly",
            width=10,
            justify="right",
        )

        self.btn_apply = tb.Button(self.filter_controls, text="تطبيق الفلتر", bootstyle="primary", command=self.apply_filters)
        self.btn_clear = tb.Button(self.filter_controls, text="مسح الفلتر", bootstyle="secondary", command=self.clear_filters)

        tb.Button(
            self.filter_actions,
            text="اشتراك جديد",
            bootstyle="success",
            command=self.open_new_subscription_dialog,
        ).pack(side="left", padx=6, ipady=6)
        tb.Button(
            self.filter_actions,
            text="تصدير التقرير",
            bootstyle="info",
            command=self.export_report,
        ).pack(side="left", padx=6, ipady=6)

        self._layout_filter_controls()

    def _layout_filter_controls(self) -> None:
        if not hasattr(self, "filter_controls"):
            return

        widgets = [
            self.lbl_start,
            self.start_date_entry,
            self.lbl_end,
            self.end_date_entry,
            self.lbl_status,
            self.status_combo,
            self.lbl_pay,
            self.pay_combo,
            self.btn_apply,
            self.btn_clear,
        ]

        for w in widgets:
            try:
                w.grid_forget()
            except Exception:
                pass
            try:
                w.pack_forget()
            except Exception:
                pass

        if hasattr(self, "_mobile_btns_row"):
            try:
                self._mobile_btns_row.grid_forget()
            except Exception:
                pass

        for i in range(20):
            try:
                self.filter_controls.columnconfigure(i, weight=0)
            except Exception:
                pass

        if getattr(self, "breakpoint", "desktop") == "mobile":
            self.filter_controls.columnconfigure(0, weight=1)
            row = 0
            pairs = [
                (self.lbl_start, self.start_date_entry),
                (self.lbl_end, self.end_date_entry),
                (self.lbl_status, self.status_combo),
                (self.lbl_pay, self.pay_combo),
            ]
            for lbl, field in pairs:
                lbl.grid(row=row, column=1, sticky="e", padx=(0, 6), pady=3)
                field.grid(row=row, column=0, sticky="ew", pady=3)
                row += 1

            if not hasattr(self, "_mobile_btns_row") or not self._mobile_btns_row.winfo_exists():
                self._mobile_btns_row = tb.Frame(self.filter_controls)
            btns = self._mobile_btns_row
            btns.grid(row=row, column=0, columnspan=2, sticky="ew", pady=(6, 0))
            try:
                self.btn_clear.pack_forget()
                self.btn_apply.pack_forget()
            except Exception:
                pass
            self.btn_apply.pack(in_=btns, side="right", padx=6)
            self.btn_clear.pack(in_=btns, side="right", padx=6)
            return

        self.filter_controls.columnconfigure(0, weight=1)
        self.lbl_start.grid(row=0, column=9, sticky="e", padx=(0, 6), pady=3)
        self.start_date_entry.grid(row=0, column=8, sticky="w", padx=(0, 12), pady=3)
        self.lbl_end.grid(row=0, column=7, sticky="e", padx=(0, 6), pady=3)
        self.end_date_entry.grid(row=0, column=6, sticky="w", padx=(0, 12), pady=3)
        self.lbl_status.grid(row=0, column=5, sticky="e", padx=(0, 6), pady=3)
        self.status_combo.grid(row=0, column=4, sticky="w", padx=(0, 12), pady=3)
        self.lbl_pay.grid(row=0, column=3, sticky="e", padx=(0, 6), pady=3)
        self.pay_combo.grid(row=0, column=2, sticky="w", padx=(0, 12), pady=3)

        self.btn_apply.grid(row=0, column=1, sticky="w", padx=6, pady=3)
        self.btn_clear.grid(row=0, column=0, sticky="w", padx=6, pady=3)

    def create_table(self) -> None:
        wrap = tb.Frame(self)
        wrap.pack(fill="both", expand=True)
        self.table_wrap = wrap

        columns = (
            "sub_id",
            "member_name",
            "phone",
            "member_status",
            "plan",
            "start_date",
            "end_date",
            "total_amount",
            "paid_amount",
            "remaining",
            "sub_status",
            "pay_status",
        )

        self.tree = ttk.Treeview(wrap, columns=columns, show="headings", selectmode="browse")

        try:
            style = ttk.Style()
            style.configure("Subscriptions.Treeview", font=("Cairo", 10), rowheight=26)
            style.configure("Subscriptions.Treeview.Heading", font=("Cairo", 10, "bold"))
            self.tree.configure(style="Subscriptions.Treeview")
        except Exception:
            pass

        headers = {
            "sub_id": ("#", 55, "center"),
            "member_name": ("اسم العضو", 200, "e"),
            "phone": ("رقم الهاتف", 115, "center"),
            "member_status": ("حالة العضو", 95, "center"),
            "plan": ("نوع الباقة", 85, "center"),
            "start_date": ("تاريخ البداية", 100, "center"),
            "end_date": ("تاريخ الانتهاء", 100, "center"),
            "total_amount": ("الإجمالي", 110, "center"),
            "paid_amount": ("المدفوع", 110, "center"),
            "remaining": ("المتبقي", 110, "center"),
            "sub_status": ("حالة الاشتراك", 95, "center"),
            "pay_status": ("حالة الدفع", 85, "center"),
        }

        for col, (txt, w, anchor) in headers.items():
            self.tree.heading(col, text=txt)
            stretch = col != "sub_id"
            self.tree.column(col, width=w, minwidth=w, anchor=anchor, stretch=stretch)

        self.tree_vsb = ttk.Scrollbar(wrap, orient="vertical", command=self.tree.yview)
        self.tree_hsb = ttk.Scrollbar(wrap, orient="horizontal", command=self.tree.xview)
        self.tree.configure(yscrollcommand=self.tree_vsb.set, xscrollcommand=self.tree_hsb.set)

        self.tree.pack(side="left", fill="both", expand=True)
        self.tree_vsb.pack(side="right", fill="y")
        self.tree_hsb.pack(side="bottom", fill="x")

        self.cards_canvas = tk.Canvas(wrap, highlightthickness=0)
        self.cards_vsb = ttk.Scrollbar(wrap, orient="vertical", command=self.cards_canvas.yview)
        self.cards_canvas.configure(yscrollcommand=self.cards_vsb.set)

        self.cards_inner = tb.Frame(self.cards_canvas)
        self.cards_window = self.cards_canvas.create_window((0, 0), window=self.cards_inner, anchor="nw")

        self.cards_inner.bind(
            "<Configure>",
            lambda _e: self.cards_canvas.configure(scrollregion=self.cards_canvas.bbox("all")),
        )
        self.cards_canvas.bind(
            "<Configure>",
            lambda e: self.cards_canvas.itemconfigure(self.cards_window, width=e.width),
        )

        self.tree.tag_configure("active", background="#ffffff")
        self.tree.tag_configure("expired", background="#fee2e2")
        self.tree.tag_configure("expiring", background="#fef3c7")
        self.tree.tag_configure("cancelled", background="#e5e7eb")
        self.tree.tag_configure("frozen", background="#e0f2fe")
        self.tree.tag_configure("member_inactive", background="#fca5a5", foreground="#7f1d1d")

        self.tree.bind("<Double-1>", lambda _e: self.view_subscription())
        self.tree.bind("<Return>", lambda _e: self.view_subscription())

        self.context_menu = tk.Menu(self, tearoff=0)
        self.context_menu.add_command(label="عرض", command=self.view_subscription)
        self.context_menu.add_command(label="تعديل", command=self.edit_subscription)
        self.context_menu.add_command(label="تجديد", command=self.open_renewal_dialog)
        self.context_menu.add_separator()
        self.context_menu.add_command(label="إلغاء", command=self.cancel_subscription)
        self.context_menu.add_command(label="طباعة إيصال", command=self.print_receipt)

        self.tree.bind("<Button-3>", self._show_context_menu)

        self._apply_responsive_layout()

    def _apply_responsive_layout(self) -> None:
        try:
            if self.breakpoint == "mobile":
                try:
                    self.tree.pack_forget()
                    self.tree_vsb.pack_forget()
                    self.tree_hsb.pack_forget()
                except Exception:
                    pass
                self.cards_canvas.pack(side="left", fill="both", expand=True)
                self.cards_vsb.pack(side="right", fill="y")
            else:
                try:
                    self.cards_canvas.pack_forget()
                    self.cards_vsb.pack_forget()
                except Exception:
                    pass
                self.tree.pack(side="left", fill="both", expand=True)
                self.tree_vsb.pack(side="right", fill="y")
                self.tree_hsb.pack(side="bottom", fill="x")
        except Exception:
            pass

    def create_quick_actions(self) -> None:
        panel = tb.Labelframe(self, text="إجراءات سريعة", padding=12)
        panel.pack(fill="x", pady=(12, 0))

        tb.Button(panel, text="🔄 تحديث", bootstyle="secondary", command=self.refresh).pack(side="right", padx=6)
        tb.Button(panel, text="⏳ الاشتراكات المنتهية", bootstyle="warning", command=self.show_expired_only).pack(
            side="right", padx=6
        )
        tb.Button(panel, text="✅ الاشتراكات النشطة", bootstyle="success", command=self.show_active_only).pack(
            side="right", padx=6
        )

        self.health_meter: object | None = None
        try:
            self.health_meter = Meter(
                panel,
                metersize=110,
                padding=6,
                amounttotal=100,
                amountused=0,
                metertype="semi",
                subtext="نسبة التجديد",
                bootstyle="info",
            )
            self.health_meter.pack(side="left")  # type: ignore[attr-defined]
        except Exception:
            try:
                wrap = tb.Frame(panel)
                wrap.pack(side="left")
                tb.Label(wrap, text="نسبة التجديد", font=FONTS["small"], anchor="e").pack(fill="x")
                self.health_meter = tb.Progressbar(wrap, maximum=100, value=0, length=120, bootstyle="info")
                self.health_meter.pack(fill="x", pady=(4, 0))
            except Exception:
                self.health_meter = None

    def _set_health_meter_value(self, pct: int) -> None:
        widget = getattr(self, "health_meter", None)
        if widget is None:
            return
        try:
            widget.configure(amountused=int(pct))  # type: ignore[attr-defined]
            return
        except Exception:
            pass
        try:
            widget.configure(value=int(pct))  # type: ignore[attr-defined]
            return
        except Exception:
            pass
        try:
            widget["value"] = int(pct)  # type: ignore[index]
        except Exception:
            pass

    # ------------------------------
    # Data loading
    # ------------------------------

    def load_summary_cards(self) -> None:
        """Load counts/sums for summary cards."""

        if self.db is None:
            return

        today = _today_str()
        with self.db.get_connection() as conn:
            cur = conn.cursor()

            active = cur.execute(
                """
                SELECT COUNT(*) AS c
                FROM subscriptions
                WHERE status = 'active' AND date(end_date) >= date(?)
                """,
                (today,),
            ).fetchone()["c"]

            expiring = cur.execute(
                """
                SELECT COUNT(*) AS c
                FROM subscriptions
                WHERE status = 'active' AND date(end_date) BETWEEN date(?) AND date(?, '+7 days')
                """,
                (today, today),
            ).fetchone()["c"]

            expired = cur.execute(
                """
                SELECT COUNT(*) AS c
                FROM subscriptions
                WHERE date(end_date) < date(?) OR status = 'expired'
                """,
                (today,),
            ).fetchone()["c"]

            ms = _month_start(date.today()).isoformat()
            revenue = cur.execute(
                """
                SELECT COALESCE(SUM(amount), 0) AS total
                FROM payments
                WHERE date(payment_date) >= date(?)
                """,
                (ms,),
            ).fetchone()["total"]

        self.summary_vars["active"].set(str(int(active)))
        self.summary_vars["expiring"].set(str(int(expiring)))
        self.summary_vars["expired"].set(str(int(expired)))
        self.summary_vars["revenue"].set(format_money(float(revenue), db=self.db, decimals=0))

        # A simple renewal "health" indicator
        try:
            total_active = int(active)
            total_exp = int(expiring)
            used = 0 if total_active == 0 else int(max(0, 100 - (total_exp / total_active) * 100))
            self._set_health_meter_value(used)
        except Exception:
            pass

    def load_subscriptions_data(self, filters: dict[str, Any] | None = None) -> None:
        """Load subscriptions with member and plan info."""

        if self.db is None:
            return

        where = []
        params: list[Any] = []

        if filters:
            if filters.get("start"):
                where.append("date(s.start_date) >= date(?)")
                params.append(filters["start"])
            if filters.get("end"):
                where.append("date(s.end_date) <= date(?)")
                params.append(filters["end"])
            if filters.get("status") and filters["status"] != "الكل":
                status_map = {
                    "نشط": "active",
                    "منتهي": "expired",
                    "معلق": "frozen",
                    "ملغي": "cancelled",
                }
                where.append("s.status = ?")
                params.append(status_map.get(filters["status"], filters["status"]))

        sql = """
            SELECT s.id AS sub_id,
                   m.id AS member_id,
                   m.first_name,
                   m.last_name,
                   m.phone,
                   m.status AS member_status,
                   st.name_ar AS plan_name,
                   st.duration_months,
                   st.price AS plan_price,
                   s.start_date,
                   s.end_date,
                   s.amount_paid,
                   s.status,
                   s.payment_method,
                   s.notes
            FROM subscriptions s
            JOIN members m ON m.id = s.member_id
            JOIN subscription_types st ON st.id = s.subscription_type_id
        """

        if where:
            sql += " WHERE " + " AND ".join(where)

        sql += " ORDER BY date(s.created_at) DESC, s.id DESC"

        with self.db.get_connection() as conn:
            rows = conn.execute(sql, tuple(params)).fetchall()
            self._rows = [dict(r) for r in rows]

        # Apply payment status filter in-memory (based on price vs amount_paid)
        if filters and filters.get("payment_status") and filters["payment_status"] != "الكل":
            need = filters["payment_status"]
            self._rows = [r for r in self._rows if self._payment_status_ar(r) == need]

        self._render_table()
        self._render_cards()

    def _render_table(self) -> None:
        for iid in self.tree.get_children():
            self.tree.delete(iid)

        today = date.today()

        for r in self._rows:
            name = f"{r.get('first_name', '')} {r.get('last_name', '')}".strip()
            phone = str(r.get("phone") or "")
            plan = str(r.get("plan_name") or "")

            start_date = str(r.get("start_date") or "")
            end_date = str(r.get("end_date") or "")

            total = float(r.get("plan_price") or 0)
            paid = float(r.get("amount_paid") or 0)
            remaining = max(0.0, total - paid)

            sub_status_ar, tag = self._subscription_status_ar_and_tag(r, today)
            pay_status_ar = self._payment_status_ar(r)

            ms_raw = str(r.get("member_status") or "active")
            ms_key = ms_raw.strip().lower()
            is_active_member = ms_key == "active"
            ms_ar = {"active": "نشط", "inactive": "غير نشط", "frozen": "مجمد"}.get(ms_key, ms_raw)

            if not is_active_member:
                sub_status_ar = "غير نشط"

            if not is_active_member:
                tags = ["member_inactive"]
            else:
                tags = [tag]

            self.tree.insert(
                "",
                "end",
                iid=str(r.get("sub_id")),
                values=(
                    r.get("sub_id"),
                    name,
                    phone,
                    ms_ar,
                    plan,
                    start_date,
                    end_date,
                    format_money(float(total), db=self.db, decimals=0),
                    format_money(float(paid), db=self.db, decimals=0),
                    format_money(float(remaining), db=self.db, decimals=0),
                    sub_status_ar,
                    pay_status_ar,
                ),
                tags=tuple(tags),
            )

    def _render_cards(self) -> None:
        if getattr(self, "breakpoint", "desktop") != "mobile":
            return
        if not hasattr(self, "cards_inner"):
            return

        for child in self.cards_inner.winfo_children():
            child.destroy()

        today = date.today()
        for r in self._rows:
            sub_id = int(r.get("sub_id") or 0)
            name = f"{r.get('first_name', '')} {r.get('last_name', '')}".strip() or "-"
            phone = str(r.get("phone") or "-")
            plan = str(r.get("plan_name") or "-")
            start_date = str(r.get("start_date") or "-")
            end_date = str(r.get("end_date") or "-")

            total = float(r.get("plan_price") or 0)
            paid = float(r.get("amount_paid") or 0)
            remaining = total - paid
            sub_status_ar, _tag = self._subscription_status_ar_and_tag(r, today)
            pay_status_ar = self._payment_status_ar(r)

            member_status = str(r.get("member_status") or "active")
            card_style = "danger" if member_status != "active" else "secondary"
            if str(member_status).strip().lower() != "active":
                sub_status_ar = "غير نشط"
            card = tb.Frame(self.cards_inner, padding=10, bootstyle=card_style)
            card.pack(fill="x", pady=6)
            card.configure(cursor="hand2")

            top = tb.Frame(card)
            top.pack(fill="x")
            tb.Label(top, text=f"#{sub_id}", font=("Cairo", 11, "bold"), anchor="e").pack(side="right")
            tb.Label(top, text=sub_status_ar, font=("Cairo", 10, "bold"), anchor="w").pack(side="left")

            tb.Label(card, text=name, font=("Cairo", 13, "bold"), anchor="e").pack(fill="x", pady=(6, 0))
            tb.Label(card, text=f"هاتف: {phone}", font=FONTS["small"], anchor="e").pack(fill="x")
            tb.Label(card, text=f"الباقة: {plan}", font=FONTS["small"], anchor="e").pack(fill="x")
            tb.Label(card, text=f"من: {start_date}  إلى: {end_date}", font=FONTS["small"], anchor="e").pack(fill="x")
            tb.Label(
                card,
                text=f"الإجمالي: {total:,.0f} | المدفوع: {paid:,.0f} | المتبقي: {remaining:,.0f} | الدفع: {pay_status_ar}",
                font=FONTS["small"],
                anchor="e",
                justify="right",
            ).pack(fill="x", pady=(4, 0))

            def open_sub(_e=None, sid=sub_id):
                self._selected_sub_id = int(sid)
                try:
                    self.tree.selection_set(str(sid))
                except Exception:
                    pass
                self.view_subscription()

            for w in (card, top):
                w.bind("<Button-1>", open_sub)

    def _subscription_status_ar_and_tag(self, r: dict[str, Any], today: date) -> tuple[str, str]:
        status = str(r.get("status") or "active")
        end_date = str(r.get("end_date") or "")

        # Compute expiration
        days_left = None
        try:
            d_end = datetime.strptime(end_date, config.DATE_FORMAT).date()
            days_left = (d_end - today).days
        except Exception:
            days_left = None

        if status == "cancelled":
            return "ملغي", "cancelled"
        if status == "frozen":
            return "معلق", "frozen"

        if days_left is not None and days_left < 0:
            return "منتهي", "expired"
        if days_left is not None and days_left <= 7 and status == "active":
            return "نشط", "expiring"

        return ("نشط" if status == "active" else status), ("active" if status == "active" else "active")

    def _payment_status_ar(self, r: dict[str, Any]) -> str:
        total = float(r.get("plan_price") or 0)
        paid = float(r.get("amount_paid") or 0)
        if paid <= 0:
            return "غير مدفوع"
        if paid + 1e-9 >= total:
            return "مدفوع"
        return "جزئي"

    # ------------------------------
    # Filters
    # ------------------------------

    def apply_filters(self) -> None:
        filters = {
            "start": self._safe_date_from_entry(self.start_date_entry),
            "end": self._safe_date_from_entry(self.end_date_entry),
            "status": self.status_var.get(),
            "payment_status": self.payment_status_var.get(),
        }
        self.load_subscriptions_data(filters=filters)
        self.load_summary_cards()

    def clear_filters(self) -> None:
        self.status_var.set("الكل")
        self.payment_status_var.set("الكل")
        try:
            self.start_date_entry.entry.delete(0, tk.END)
            self.end_date_entry.entry.delete(0, tk.END)
        except Exception:
            pass
        self.load_subscriptions_data(filters=None)
        self.load_summary_cards()

    def _safe_date_from_entry(self, entry: DateEntry) -> str | None:
        try:
            # DateEntry stores a date object or string
            val = entry.date
            if isinstance(val, date):
                return val.isoformat()
        except Exception:
            pass

        try:
            raw = entry.entry.get().strip()
            if not raw:
                return None
            # Accept ISO
            datetime.strptime(raw, config.DATE_FORMAT)
            return raw
        except Exception:
            return None

    # ------------------------------
    # Context menu actions
    # ------------------------------

    def _show_context_menu(self, event) -> None:
        try:
            iid = self.tree.identify_row(event.y)
            if iid:
                self.tree.selection_set(iid)
            self.context_menu.tk_popup(event.x_root, event.y_root)
        finally:
            try:
                self.context_menu.grab_release()
            except Exception:
                pass

    def _selected_subscription_id(self) -> int | None:
        sel = self.tree.selection()
        if not sel:
            if self._selected_sub_id is not None:
                return self._selected_sub_id
            return None
        values = self.tree.item(sel[0], "values")
        try:
            return int(values[0])
        except Exception:
            return None

    def view_subscription(self) -> None:
        sub_id = self._selected_subscription_id()
        if sub_id is None or self.db is None:
            return

        try:
            sub = self.db.get_subscription_by_id(sub_id)
        except Exception as e:
            messagebox.showerror("خطأ", f"تعذر تحميل الاشتراك: {e}")
            return

        if not sub:
            messagebox.showwarning("تنبيه", "الاشتراك غير موجود")
            return

        ReceiptDialog(self.winfo_toplevel(), sub, title="تفاصيل الاشتراك").wait_window()

    def edit_subscription(self) -> None:
        # Minimal implementation: open renewal dialog (safe) as "edit".
        self.open_renewal_dialog()

    def open_new_subscription_dialog(self) -> None:
        if self.db is None:
            messagebox.showerror("خطأ", "قاعدة البيانات غير جاهزة")
            return

        dlg = NewSubscriptionDialog(self.winfo_toplevel(), self.db, user_data=self.user_data)
        self.wait_window(dlg)
        if dlg.saved:
            self.refresh()

    def open_renewal_dialog(self) -> None:
        if self.db is None:
            messagebox.showerror("خطأ", "قاعدة البيانات غير جاهزة")
            return

        sub_id = self._selected_subscription_id()
        if sub_id is None:
            messagebox.showwarning("تنبيه", "يرجى اختيار اشتراك")
            return

        try:
            sub = self.db.get_subscription_by_id(sub_id)
        except Exception as e:
            messagebox.showerror("خطأ", f"تعذر تحميل الاشتراك: {e}")
            return

        if not sub:
            messagebox.showwarning("تنبيه", "الاشتراك غير موجود")
            return

        dlg = RenewalDialog(self.winfo_toplevel(), self.db, sub)
        self.wait_window(dlg)
        if dlg.saved:
            self.refresh()

    def cancel_subscription(self) -> None:
        if self.db is None:
            messagebox.showerror("خطأ", "قاعدة البيانات غير جاهزة")
            return

        sub_id = self._selected_subscription_id()
        if sub_id is None:
            messagebox.showwarning("تنبيه", "يرجى اختيار اشتراك")
            return

        if not messagebox.askyesno("تأكيد", "هل تريد إلغاء الاشتراك المحدد؟"):
            return

        ok, msg = self.db.cancel_subscription(sub_id)
        if ok:
            messagebox.showinfo("تم", "تم إلغاء الاشتراك")
            self.refresh()
        else:
            messagebox.showerror("خطأ", msg)

    def print_receipt(self) -> None:
        if self.db is None:
            return

        sub_id = self._selected_subscription_id()
        if sub_id is None:
            messagebox.showwarning("تنبيه", "يرجى اختيار اشتراك")
            return

        try:
            sub = self.db.get_subscription_by_id(sub_id)
        except Exception as e:
            messagebox.showerror("خطأ", f"تعذر تحميل بيانات الإيصال: {e}")
            return

        if not sub:
            messagebox.showwarning("تنبيه", "الاشتراك غير موجود")
            return

        ReceiptDialog(self.winfo_toplevel(), sub, title="إيصال الدفع").wait_window()

    # ------------------------------
    # Export
    # ------------------------------

    def export_report(self) -> None:
        if not self._rows:
            messagebox.showwarning("تنبيه", "لا توجد بيانات للتصدير")
            return

        from tkinter import filedialog

        file_path = filedialog.asksaveasfilename(
            title="تصدير التقرير",
            defaultextension=".xlsx",
            filetypes=[("Excel", "*.xlsx"), ("CSV", "*.csv")],
        )
        if not file_path:
            return

        try:
            if file_path.lower().endswith(".xlsx"):
                self._export_xlsx(file_path)
            else:
                self._export_csv(file_path)
            messagebox.showinfo("تم", "تم التصدير بنجاح")
        except Exception as e:
            messagebox.showerror("خطأ", f"فشل التصدير: {e}")

    def _export_csv(self, file_path: str) -> None:
        import csv

        headers = [
            "رقم الاشتراك",
            "اسم العضو",
            "رقم الهاتف",
            "نوع الباقة",
            "تاريخ البداية",
            "تاريخ الانتهاء",
            "المبلغ الكلي",
            "المدفوع",
            "المتبقي",
            "حالة الاشتراك",
            "حالة الدفع",
        ]

        with open(file_path, "w", newline="", encoding="utf-8") as f:
            w = csv.writer(f)
            w.writerow(headers)
            for r in self._rows:
                total = float(r.get("plan_price") or 0)
                paid = float(r.get("amount_paid") or 0)
                remaining = total - paid
                sub_status_ar, _tag = self._subscription_status_ar_and_tag(r, date.today())
                pay_status_ar = self._payment_status_ar(r)

                w.writerow(
                    [
                        r.get("sub_id"),
                        f"{r.get('first_name', '')} {r.get('last_name', '')}".strip(),
                        r.get("phone", ""),
                        r.get("plan_name", ""),
                        r.get("start_date", ""),
                        r.get("end_date", ""),
                        total,
                        paid,
                        remaining,
                        sub_status_ar,
                        pay_status_ar,
                    ]
                )

            # Summary row
            w.writerow([])
            w.writerow(["ملخص", "", "", "", "", "", "", "", "", "", ""])
            w.writerow(["الاشتراكات النشطة", self.summary_vars["active"].get()])
            w.writerow(["تنتهي خلال 7 أيام", self.summary_vars["expiring"].get()])
            w.writerow(["اشتراكات منتهية", self.summary_vars["expired"].get()])
            w.writerow(["الإيرادات الشهرية", self.summary_vars["revenue"].get()])

    def _export_xlsx(self, file_path: str) -> None:
        try:
            import openpyxl  # type: ignore
            from openpyxl.styles import Alignment, Font  # type: ignore
        except Exception as e:
            raise RuntimeError("openpyxl غير مثبت. اختر CSV أو قم بتثبيت openpyxl") from e

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Subscriptions"

        headers = [
            "رقم الاشتراك",
            "اسم العضو",
            "رقم الهاتف",
            "نوع الباقة",
            "تاريخ البداية",
            "تاريخ الانتهاء",
            "المبلغ الكلي",
            "المدفوع",
            "المتبقي",
            "حالة الاشتراك",
            "حالة الدفع",
        ]
        ws.append(headers)

        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="right")

        for r in self._rows:
            total = float(r.get("plan_price") or 0)
            paid = float(r.get("amount_paid") or 0)
            remaining = total - paid
            sub_status_ar, _tag = self._subscription_status_ar_and_tag(r, date.today())
            pay_status_ar = self._payment_status_ar(r)

            ws.append(
                [
                    r.get("sub_id"),
                    f"{r.get('first_name', '')} {r.get('last_name', '')}".strip(),
                    r.get("phone", ""),
                    r.get("plan_name", ""),
                    r.get("start_date", ""),
                    r.get("end_date", ""),
                    total,
                    paid,
                    remaining,
                    sub_status_ar,
                    pay_status_ar,
                ]
            )

        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(horizontal="right")

        # Summary
        ws.append([])
        ws.append(["ملخص", "", "", "", "", "", "", "", "", "", ""])
        ws.append(["الاشتراكات النشطة", self.summary_vars["active"].get()])
        ws.append(["تنتهي خلال 7 أيام", self.summary_vars["expiring"].get()])
        ws.append(["اشتراكات منتهية", self.summary_vars["expired"].get()])
        ws.append(["الإيرادات الشهرية", self.summary_vars["revenue"].get()])

        wb.save(file_path)

    # ------------------------------
    # Quick actions
    # ------------------------------

    def refresh(self) -> None:
        self.load_summary_cards()
        self.load_subscriptions_data()

    def show_active_only(self) -> None:
        self.status_var.set("نشط")
        self.apply_filters()

    def show_expired_only(self) -> None:
        self.status_var.set("منتهي")
        self.apply_filters()

    # ------------------------------
    # Notifications
    # ------------------------------

    def check_expiring_subscriptions(self) -> None:
        """Show a toast-like message if there are subscriptions expiring soon."""

        if self.db is None:
            return

        try:
            exp = self.db.get_expiring_subscriptions(days=7)
            if exp:
                Messagebox.ok(
                    title="تنبيه",
                    message=f"يوجد {len(exp)} اشتراك/اشتراكات تنتهي خلال 7 أيام.",
                    parent=self.winfo_toplevel(),
                )
        except Exception:
            pass


class NewSubscriptionDialog(tk.Toplevel):
    """Dialog for creating a new subscription."""

    def __init__(self, parent: tk.Misc, db: DatabaseManager, user_data: dict | None = None) -> None:
        super().__init__(parent)
        self.db = db
        self.user_data = user_data or {}
        self.saved: bool = False
        self._saving: bool = False

        self.title("اشتراك جديد")
        self.geometry("760x620")
        self.minsize(420, 520)
        self.resizable(True, True)
        self.grab_set()

        self.member_search_var = tk.StringVar(master=self, value="")
        self.selected_member: dict[str, Any] | None = None

        self.plan_var = tk.StringVar(master=self, value="")
        self.start_var = tk.StringVar(master=self, value=date.today().isoformat())
        self.end_var = tk.StringVar(master=self, value="")

        self.total_var = tk.StringVar(master=self, value="0")
        self.discount_var = tk.StringVar(master=self, value="0")
        self.final_var = tk.StringVar(master=self, value="0")
        self.paid_var = tk.StringVar(master=self, value="0")
        self.remaining_var = tk.StringVar(master=self, value="0")
        self.pay_method_var = tk.StringVar(master=self, value="cash")

        self._plans: list[dict[str, Any]] = []

        self.create_widgets()
        self._load_plans()
        self._center()

        self.member_search_var.trace_add("write", lambda *_: self._search_members())
        self.discount_var.trace_add("write", lambda *_: self._recalc_amounts())
        self.paid_var.trace_add("write", lambda *_: self._recalc_amounts())

    def create_widgets(self) -> None:
        container = tb.Frame(self, padding=14)
        container.pack(side="top", fill="both", expand=True)

        # Buttons (top-right)
        btns = tb.Frame(container)
        btns.pack(fill="x", pady=(0, 10))

        self.btn_save = tb.Button(btns, text="حفظ الاشتراك", bootstyle="success", command=self._save)
        self.btn_save.pack(side="right", padx=6)
        self.btn_save_print = tb.Button(btns, text="حفظ وطباعة الإيصال", bootstyle="info", command=lambda: self._save(print_after=True))
        self.btn_save_print.pack(side="right", padx=6)
        tb.Button(btns, text="إلغاء", bootstyle="secondary", command=self.destroy).pack(side="left")

        # Member selection
        mem = tb.Labelframe(container, text="اختيار العضو", padding=12)
        mem.pack(fill="x")

        tb.Label(mem, text="البحث عن عضو بالاسم أو الهاتف", font=("Cairo", 10, "bold"), anchor="e").pack(fill="x")
        self.member_entry = tb.Entry(mem, textvariable=self.member_search_var, justify="right")
        self.member_entry.pack(fill="x", pady=(6, 8), ipady=5)

        self.results = tk.Listbox(mem, height=5)
        self.results.pack(fill="x")
        self.results.bind("<<ListboxSelect>>", lambda _e: self._on_member_select())

        self.selected_label = tb.Label(mem, text="لم يتم اختيار عضو", font=FONTS["small"], foreground=COLORS["text_light"], anchor="e")
        self.selected_label.pack(fill="x", pady=(8, 0))

        # Plan selection
        plan = tb.Labelframe(container, text="اختيار الباقة", padding=12)
        plan.pack(fill="x", pady=10)

        self.plan_combo = tb.Combobox(plan, textvariable=self.plan_var, state="readonly", justify="right")
        self.plan_combo.pack(fill="x")
        self.plan_combo.bind("<<ComboboxSelected>>", lambda _e: self._on_plan_change())

        # Dates
        dates = tb.Labelframe(container, text="التواريخ", padding=12)
        dates.pack(fill="x")

        row = tb.Frame(dates)
        row.pack(fill="x")

        tb.Label(row, text="تاريخ البداية", font=("Cairo", 10, "bold")).pack(side="right", padx=(0, 6))
        self.start_entry = DateEntry(row, width=12, bootstyle="info", dateformat="%Y-%m-%d")
        self.start_entry.pack(side="right", padx=(0, 14))

        tb.Label(row, text="تاريخ الانتهاء", font=("Cairo", 10, "bold")).pack(side="right", padx=(0, 6))
        self.end_entry = DateEntry(row, width=12, bootstyle="info", dateformat="%Y-%m-%d")
        self.end_entry.pack(side="right")

        # Initialize and sync date variables
        try:
            self.start_entry.date = date.today()
            self.start_var.set(date.today().isoformat())
        except Exception:
            pass

        def _sync_dates_from_widgets(_e=None) -> None:
            try:
                s = self.start_entry.date
                if isinstance(s, date):
                    self.start_var.set(s.isoformat())
            except Exception:
                pass
            try:
                e = self.end_entry.date
                if isinstance(e, date):
                    self.end_var.set(e.isoformat())
            except Exception:
                pass

        for ev in ("<<DateEntrySelected>>", "<<CalendarSelected>>", "<FocusOut>"):
            try:
                self.start_entry.bind(ev, _sync_dates_from_widgets)
            except Exception:
                pass
            try:
                self.end_entry.bind(ev, _sync_dates_from_widgets)
            except Exception:
                pass

        self.duration_label = tb.Label(dates, text="مدة الاشتراك: -", font=FONTS["small"], anchor="e")
        self.duration_label.pack(fill="x", pady=(8, 0))

        # Payment
        pay = tb.Labelframe(container, text="بيانات الدفع", padding=12)
        pay.pack(fill="x", pady=10)

        grid = tb.Frame(pay)
        grid.pack(fill="x")

        def field(grow: int, label: str, var: tk.StringVar) -> None:
            fr = tb.Frame(grid)
            fr.grid(row=grow, column=0, sticky="ew", pady=4)
            fr.columnconfigure(0, weight=1)
            tb.Label(fr, text=label, font=("Cairo", 10, "bold"), anchor="e").pack(side="right")
            tb.Entry(fr, textvariable=var, justify="right").pack(side="left", fill="x", expand=True)

        field(0, "المبلغ الكلي", self.total_var)
        field(1, "الخصم", self.discount_var)
        field(2, "المبلغ بعد الخصم", self.final_var)
        field(3, "المبلغ المدفوع", self.paid_var)
        field(4, "المتبقي", self.remaining_var)

        pm_row = tb.Frame(grid)
        pm_row.grid(row=5, column=0, sticky="ew", pady=4)
        pm_row.columnconfigure(0, weight=1)
        tb.Label(pm_row, text="طريقة الدفع", font=("Cairo", 10, "bold"), anchor="e").pack(side="right")
        pm = tb.Combobox(
            pm_row,
            textvariable=self.pay_method_var,
            values=list(config.PAYMENT_METHODS.keys()),
            state="readonly",
            justify="right",
            width=18,
        )
        pm.pack(side="left")


    def _center(self) -> None:
        self.update_idletasks()
        w = self.winfo_width()
        h = self.winfo_height()
        sw = self.winfo_screenwidth()
        sh = self.winfo_screenheight()
        x = (sw - w) // 2
        y = (sh - h) // 2
        self.geometry(f"{w}x{h}+{x}+{y}")

    def _load_plans(self) -> None:
        try:
            self._plans = self.db.get_all_subscription_types(active_only=True)
            values = []
            for p in self._plans:
                values.append(f"{p.get('name_ar', '')} - {format_money(float(p.get('price', 0) or 0), db=self.db, decimals=0)}")
            self.plan_combo.configure(values=values)
        except Exception:
            self._plans = []

    def _search_members(self) -> None:
        txt = self.member_search_var.get().strip()
        self.results.delete(0, tk.END)

        if not txt:
            return

        try:
            q = f"%{txt}%"
            with self.db.get_connection() as conn:
                rows = conn.execute(
                    """
                    SELECT id, member_code, first_name, last_name, phone
                    FROM members
                    WHERE member_code LIKE ? OR phone LIKE ? OR first_name LIKE ? OR last_name LIKE ?
                    ORDER BY id DESC
                    LIMIT 20
                    """,
                    (q, q, q, q),
                ).fetchall()

            self._member_results = [dict(r) for r in rows]
            for r in self._member_results:
                self.results.insert(
                    tk.END,
                    f"{r.get('member_code', '')} - {r.get('first_name', '')} {r.get('last_name', '')} - {r.get('phone', '')}",
                )
        except Exception:
            self._member_results = []

    def _on_member_select(self) -> None:
        try:
            idx = int(self.results.curselection()[0])
        except Exception:
            return

        self.selected_member = self._member_results[idx]
        self.selected_label.configure(
            text=f"العضو المختار: {self.selected_member.get('first_name', '')} {self.selected_member.get('last_name', '')} - {self.selected_member.get('phone', '')}",
            foreground=COLORS["text"],
        )

    def _on_plan_change(self) -> None:
        plan = self._selected_plan()
        if not plan:
            return

        total = float(plan.get("price") or 0)
        self.total_var.set(f"{total:.0f}")

        try:
            start = self.start_entry.date
            if isinstance(start, date):
                end = _add_months(start, int(plan.get("duration_months") or 1))
                self.end_entry.date = end
                self.duration_label.configure(text=f"مدة الاشتراك: {int(plan.get('duration_months') or 1)} شهر")
        except Exception:
            pass

        self._recalc_amounts()

    def _selected_plan(self) -> dict[str, Any] | None:
        val = self.plan_var.get().strip()
        if not val:
            return None
        name = val.split("-")[0].strip()
        for p in self._plans:
            if str(p.get("name_ar", "")).strip() == name:
                return p
        return None

    def _recalc_amounts(self) -> None:
        try:
            total = float(self.total_var.get() or 0)
        except Exception:
            total = 0

        try:
            discount = float(self.discount_var.get() or 0)
        except Exception:
            discount = 0

        discount = max(0.0, min(discount, total))
        final = total - discount

        try:
            paid = float(self.paid_var.get() or 0)
        except Exception:
            paid = 0

        paid = max(0.0, paid)
        remaining = final - paid

        self.final_var.set(f"{final:.0f}")
        self.remaining_var.set(f"{remaining:.0f}")

    def _coerce_date(self, v: object) -> date:
        if isinstance(v, date):
            return v
        if isinstance(v, str):
            txt = v.strip()
            for fmt in ("%Y-%m-%d", "%m/%d/%y", "%m/%d/%Y", "%d/%m/%Y"):
                try:
                    return datetime.strptime(txt, fmt).date()
                except Exception:
                    continue
        raise ValueError("invalid date")

    def _save(self, print_after: bool = False) -> None:
        if getattr(self, "_saving", False):
            return
        self._saving = True
        try:
            self.btn_save.configure(state="disabled")
            self.btn_save_print.configure(state="disabled")
        except Exception:
            pass

        if self.selected_member is None:
            messagebox.showwarning("تنبيه", "يرجى اختيار عضو")
            try:
                self.btn_save.configure(state="normal")
                self.btn_save_print.configure(state="normal")
            except Exception:
                pass
            self._saving = False
            return

        plan = self._selected_plan()
        if not plan:
            messagebox.showwarning("تنبيه", "يرجى اختيار باقة")
            try:
                self.btn_save.configure(state="normal")
                self.btn_save_print.configure(state="normal")
            except Exception:
                pass
            self._saving = False
            return

        try:
            start = self._coerce_date(getattr(self.start_entry, "date", None))
            end = self._coerce_date(getattr(self.end_entry, "date", None))
            self.start_var.set(start.isoformat())
            self.end_var.set(end.isoformat())
        except Exception:
            messagebox.showerror("خطأ", "تواريخ غير صحيحة")
            try:
                self.btn_save.configure(state="normal")
                self.btn_save_print.configure(state="normal")
            except Exception:
                pass
            self._saving = False
            return

        try:
            paid = float(self.paid_var.get() or 0)
        except Exception:
            paid = 0

        if paid < 0:
            messagebox.showerror("خطأ", "المبلغ المدفوع غير صحيح")
            try:
                self.btn_save.configure(state="normal")
                self.btn_save_print.configure(state="normal")
            except Exception:
                pass
            self._saving = False
            return

        # Note: current DB schema stores only amount_paid; we keep discount/final in notes.
        notes = f"Discount={self.discount_var.get()} Final={self.final_var.get()} Remaining={self.remaining_var.get()}"

        ok, msg, sub_id = self.db.create_subscription(
            member_id=int(self.selected_member["id"]),
            subscription_type_id=int(plan["id"]),
            amount_paid=float(paid),
            payment_method=self.pay_method_var.get() or "cash",
            start_date=start.isoformat(),
            created_by=self.user_data.get("id") if isinstance(self.user_data, dict) else None,
        )

        if not ok or sub_id is None:
            messagebox.showerror("خطأ", msg)
            try:
                self.btn_save.configure(state="normal")
                self.btn_save_print.configure(state="normal")
            except Exception:
                pass
            self._saving = False
            return

        # Add notes if any
        try:
            with self.db.get_connection() as conn:
                conn.execute("UPDATE subscriptions SET notes = ? WHERE id = ?", (notes, sub_id))
                conn.commit()
        except Exception:
            pass

        self.saved = True
        messagebox.showinfo("تم", "تم حفظ الاشتراك بنجاح")

        if print_after:
            try:
                sub = self.db.get_subscription_by_id(sub_id)
                if sub:
                    ReceiptDialog(self, sub, title="إيصال الدفع").wait_window()
            except Exception:
                pass

        self.destroy()


class RenewalDialog(NewSubscriptionDialog):
    """Dialog for renewing a subscription (member pre-selected)."""

    def __init__(self, parent: tk.Misc, db: DatabaseManager, subscription: dict[str, Any]) -> None:
        self._existing_subscription = subscription
        super().__init__(parent, db)
        self.title("تجديد الاشتراك")

        # Preselect member
        self.selected_member = {
            "id": subscription.get("member_id"),
            "member_code": subscription.get("member_code"),
            "first_name": subscription.get("first_name"),
            "last_name": subscription.get("last_name"),
            "phone": subscription.get("phone"),
        }
        self.selected_label.configure(
            text=f"العضو: {self.selected_member.get('first_name', '')} {self.selected_member.get('last_name', '')} - {self.selected_member.get('phone', '')}",
            foreground=COLORS["text"],
        )

        # Disable member search
        try:
            self.member_entry.configure(state="disabled")
            self.results.configure(state="disabled")
        except Exception:
            pass

        # Start date defaults to day after current end date
        try:
            current_end = datetime.strptime(str(subscription.get("end_date")), config.DATE_FORMAT).date()
            next_start = current_end + timedelta(days=1)
            self.start_entry.date = next_start
        except Exception:
            pass

        # Try to preselect same plan if possible
        try:
            type_name_ar = subscription.get("type_name_ar")
            if type_name_ar:
                for i, p in enumerate(self._plans):
                    if str(p.get("name_ar")) == str(type_name_ar):
                        self.plan_combo.current(i)
                        self.plan_var.set(self.plan_combo.get())
                        self._on_plan_change()
                        break
        except Exception:
            pass


class ReceiptDialog(tk.Toplevel):
    """Simple receipt/details preview dialog."""

    def __init__(self, parent: tk.Misc, subscription: dict[str, Any], title: str = "إيصال") -> None:
        super().__init__(parent)
        self.title(title)
        self.geometry("520x520")
        self.minsize(420, 420)
        self.resizable(True, True)
        self.grab_set()

        container = tb.Frame(self, padding=14)
        container.pack(fill="both", expand=True)

        tb.Label(container, text=config.APP_NAME, font=FONTS["subheading"], anchor="e").pack(fill="x")
        tb.Label(container, text=title, font=FONTS["body"], anchor="e").pack(fill="x", pady=(2, 10))

        text = tk.Text(container, wrap="word", height=18)
        text.pack(fill="both", expand=True)

        lines = []
        lines.append(f"رقم الاشتراك: {subscription.get('id')}")
        lines.append(f"العضو: {subscription.get('member_code', '')} - {subscription.get('first_name', '')} {subscription.get('last_name', '')}")
        lines.append(f"الهاتف: {subscription.get('phone', '')}")
        lines.append(f"الباقة: {subscription.get('type_name_ar', '')}")
        lines.append(f"تاريخ البداية: {subscription.get('start_date', '')}")
        lines.append(f"تاريخ الانتهاء: {subscription.get('end_date', '')}")
        lines.append(f"المبلغ المدفوع: {subscription.get('amount_paid', '')}")
        lines.append(f"طريقة الدفع: {subscription.get('payment_method', '')}")
        lines.append(f"الحالة: {subscription.get('status', '')}")
        lines.append("")
        lines.append(f"التاريخ: {datetime.now().strftime(config.DATETIME_FORMAT)}")

        text.insert("1.0", "\n".join(lines))
        text.configure(state="disabled")

        btns = tb.Frame(container)
        btns.pack(fill="x", pady=10)
        tb.Button(btns, text="🖨️ طباعة", bootstyle="secondary", command=lambda: self._print(text)).pack(side="left")
        tb.Button(btns, text="إغلاق", bootstyle="secondary", command=self.destroy).pack(side="left", padx=6)

    def _print(self, text_widget: tk.Text) -> None:
        try:
            content = text_widget.get("1.0", "end").strip()
            print_text_windows(content, filename_prefix="subscription_receipt")
        except Exception:
            messagebox.showwarning("تنبيه", "تعذر بدء الطباعة تلقائياً")


if __name__ == "__main__":
    root = tb.Window(themename="cosmo")
    root.title("SubscriptionsFrame Test")
    db = DatabaseManager()
    frame = SubscriptionsFrame(root, db, {"id": 1, "username": "admin", "role": "admin"})
    frame.pack(fill="both", expand=True)
    root.geometry("1200x700")
    root.mainloop()
